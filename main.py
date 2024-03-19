from lib import *

import openpyxl
import pandas as pd

# AGREGA NÚMERO 
num = int(input("¿Cuál es tu número? : "))
doc = input("¿Cual es el nombre del archivo a procesar? : (xlsx)")
datos = pd.read_excel('datos-Eval-2.xlsx') #,index =0,  engine ='openpyxl'
print(datos)

""" 
# LO IMPRIME COMO LISTAS DENTRO DE LISTAS
excel_dataframe = openpyxl.load_workbook("datos-Eval-2.xlsx")
dataframe = excel_dataframe.active
data = []
for row in range (0,dataframe.max_row,1):
    row2 = [row,]
    for col in dataframe.iter_cols(0,dataframe.max_column,1):
        row2.append(col[row].value)
    data.append(row2)
print(data)
"""


        
print("-------------------------------------------------------------------------------------------")


"""

#------------------------ARBOL SIMETRICO-----------------------------------------------
arrayNum= datos.values.flatten().tolist()
nodoRaiz = nodo(arrayNum[0])


for i in range(1, len(arrayNum),1):
    agregaNodos(nodoRaiz, arrayNum[i])
printArbol(nodoRaiz)
print("---------------------------------------------------------------------------------------------------------------------------------------------------------------------")
#---------IN ORDER ----------------
print("--------\033[43m\033[31m InOrder:\033[0m-----------")
inOrderArr = []
LVR(nodoRaiz, inOrderArr)
print("\033[35m InOrder:\033[0m", end=" ")
print(inOrderArr)

#------------PRE ORDER-------------
print("--------\033[43m \033[31mPreOrderArr \033[0m-----------")
preOrderArr = []
VLF(nodoRaiz, preOrderArr)
print("\033[35m PreOrder:\033[0m", end=" ")
print(preOrderArr)

#----------- POST ORDER-------------
print("--------\033[43m\033[31m PostOrderArr \033[0m-----------")
postOrderArr = []
LRV(nodoRaiz, postOrderArr)
print("\033[35m PostOrder:\033[0m", end=" ")
print(postOrderArr)


#----------------ÁRBOL ORDENADO-----------------------------------
print("árbol odenado")

nodoRaiz = num
for i in range (0,len(arrayNum),1):
    if i == 0:
        nodoRaiz = nodo(arrayNum[i])
    else:
        nodosOrdenados(nodoRaiz,nodo(arrayNum[i]))
    pass
     
printArbol(nodoRaiz)
LVR(nodoRaiz, inOrderArr)
print(inOrderArr)

"""


